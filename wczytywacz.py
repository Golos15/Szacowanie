from wynik_liczenia import wynik_liczenia
from openpyxl import load_workbook

def wczytaj():
    wb = load_workbook('E:\\ĆWICZENIE.XLSX')
    ws = wb['Arkusz1']
    komorka_nr_uzyt_pierwszy = ws['A2']
    ws.delete_cols(komorka_nr_uzyt_pierwszy.column+1, amount=2)
    lista = []
    for row in ws.iter_rows(min_row = komorka_nr_uzyt_pierwszy.row, max_col= komorka_nr_uzyt_pierwszy.column+4, min_col = komorka_nr_uzyt_pierwszy.column, max_row = None, values_only = True):
        lista.append(wynik_liczenia(row[0], row[1], row[2], row[3], row[4]))
        #print(wynik_liczenia(row[0], row[1], row[2]))
    return lista

if __name__ == '__main__':
    wczytaj()